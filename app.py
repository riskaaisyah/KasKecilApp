import streamlit as st
import pandas as pd
from st_supabase_connection import SupabaseConnection
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import datetime
import time

# --- 1. KONFIGURASI HALAMAN ---
st.set_page_config(page_title="Aplikasi Kas Kecil", layout="wide")
st.title("💰 Aplikasi Kas Kecil")

# --- 2. KONEKSI SUPABASE ---
conn = st.connection("supabase", type=SupabaseConnection)

# --- 3. FUNGSI AMBIL DATA ---
def fetch_data():
    try:
        res = conn.table("kas_kecil").select("*").execute()
        df = pd.DataFrame(res.data)
        if df.empty:
            return pd.DataFrame(columns=["id", "uraian", "vendor", "tanggal", "jumlah"])
        df['id'] = df['id'].astype(int)
        return df.sort_values("id")
    except Exception:
        return pd.DataFrame(columns=["id", "uraian", "vendor", "tanggal", "jumlah"])

opsi_uraian = ["Jamuan Makan Dinas", "Kebutuhan Kantor", "Karcis Parkir Kendaraan Operasional", "Isi BBM Kendaraan Operasional"]

# --- 4. FORM INPUT ---
with st.form("form_kas", clear_on_submit=True):
    st.subheader("Input Data Transaksi")
    col1, col2 = st.columns(2)
    with col1:
        uraian_pilih = st.selectbox("Uraian", opsi_uraian)
        vendor = st.text_input("Nama Vendor", placeholder="Ketik vendor lalu tekan Tab...")
    with col2:
        tanggal_input = st.date_input("Tanggal", value=datetime.date.today())
        jumlah = st.number_input("Jumlah (Nominal)", min_value=0, step=1000, value=None, placeholder="Masukkan nominal...")
    submit = st.form_submit_button("Simpan ke Database")

# --- 5. LOGIKA SIMPAN ---
if submit:
    if vendor.strip() != "" and jumlah is not None and jumlah > 0:
        try:
            target_uraian = "Karcis Parkir Kendaraan Operasional"
            if uraian_pilih == target_uraian:
                bln, thn = tanggal_input.month, tanggal_input.year
                res = conn.table("kas_kecil").select("*").eq("uraian", target_uraian).execute()
                df_cek = pd.DataFrame(res.data)
                found = False
                if not df_cek.empty:
                    df_cek['t_dt'] = pd.to_datetime(df_cek['tanggal'])
                    match = df_cek[(df_cek['t_dt'].dt.month == bln) & (df_cek['t_dt'].dt.year == thn)]
                    if not match.empty:
                        id_old = int(match.iloc[0]['id'])
                        new_amt = int(match.iloc[0]['jumlah'] + jumlah)
                        conn.table("kas_kecil").delete().eq("id", id_old).execute()
                        conn.table("kas_kecil").insert({"uraian": target_uraian, "vendor": vendor, "tanggal": str(tanggal_input), "jumlah": new_amt}).execute()
                        found = True
                if not found:
                    conn.table("kas_kecil").insert({"uraian": target_uraian, "vendor": vendor, "tanggal": str(tanggal_input), "jumlah": int(jumlah)}).execute()
            else:
                conn.table("kas_kecil").insert({"uraian": uraian_pilih, "vendor": vendor, "tanggal": str(tanggal_input), "jumlah": int(jumlah)}).execute()
            st.success("Data Berhasil Disimpan!"); time.sleep(0.5); st.rerun()
        except Exception as e: st.error(f"Gagal: {e}")
    else:
        st.warning("Mohon isi Nama Vendor dan Jumlah!")

# --- 6. REKAPITULASI (AUTO-FILL GAP) ---
st.divider()
df_raw = fetch_data()
nama_bulan_id = {1:"JANUARI", 2:"FEBRUARI", 3:"MARET", 4:"APRIL", 5:"MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS", 9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER"}
LIMIT_KAS = 25000000

if not df_raw.empty:
    df_raw['tanggal_dt'] = pd.to_datetime(df_raw['tanggal'], errors='coerce')
    df_raw['Kelompok_Sheet'] = ""
    for (y, m), group in df_raw.groupby([df_raw['tanggal_dt'].dt.year, df_raw['tanggal_dt'].dt.month]):
        if pd.isna(m): continue
        batch_totals = {1: 0}
        for idx, row in group.sort_values('id').iterrows():
            assigned = False; b_idx = 1
            while not assigned:
                if b_idx not in batch_totals: batch_totals[b_idx] = 0
                if batch_totals[b_idx] + row['jumlah'] <= LIMIT_KAS:
                    batch_totals[b_idx] += row['jumlah']
                    df_raw.at[idx, 'Kelompok_Sheet'] = f"{nama_bulan_id[int(m)]} ({b_idx}) {int(y)}"
                    assigned = True
                else: b_idx += 1

    list_kelompok = sorted([k for k in df_raw['Kelompok_Sheet'].unique() if k != ""], 
                           key=lambda x: (int(x.split(' ')[-1]), list(nama_bulan_id.values()).index(x.split(' ')[0])), reverse=True)

    for g in list_kelompok:
        df_g = df_raw[df_raw['Kelompok_Sheet'] == g].copy()
        tot_g = df_g['jumlah'].sum()
        with st.expander(f" {g} | Total: Rp {tot_g:,.0f} | Sisa: Rp {LIMIT_KAS-tot_g:,.0f}".replace(",","."), expanded=True):
            df_disp = df_g.copy()
            df_disp['No'] = range(1, len(df_disp)+1)
            df_disp['Uraian_View'] = df_disp.apply(lambda x: f"{x['No']} {x['uraian']}", axis=1)
            df_disp['Tgl_View'] = df_disp.apply(lambda x: "" if x['uraian'] == "Karcis Parkir Kendaraan Operasional" else x['tanggal'], axis=1)
            df_edit = df_disp[['id', 'No', 'Uraian_View', 'vendor', 'Tgl_View', 'jumlah']].copy()
            df_edit.columns = ['id', 'No', 'Uraian', 'Vendor', 'Tanggal', 'Jumlah']
            edited = st.data_editor(df_edit, key=f"ed_{g}", num_rows="dynamic", use_container_width=True,
                                    column_config={"id": None, "No": st.column_config.NumberColumn(disabled=True), "Jumlah": st.column_config.NumberColumn(format="Rp %d")})
            if st.button(f"Simpan Perubahan {g}", key=f"btn_{g}"):
                try:
                    ids_old = set(df_edit['id'].tolist()); ids_new = set(edited['id'].dropna().astype(int).tolist())
                    for d_id in (ids_old - ids_new): conn.table("kas_kecil").delete().eq("id", d_id).execute()
                    for _, row in edited.iterrows():
                        if pd.notna(row['id']):
                            u = row['Uraian'].split(' ', 1)[-1] if ' ' in row['Uraian'] else row['Uraian']
                            conn.table("kas_kecil").update({"uraian":u, "vendor":row['Vendor'], "tanggal":str(row['Tanggal']), "jumlah":int(row['Jumlah'])}).eq("id", int(row['id'])).execute()
                    st.success("Berhasil!"); time.sleep(1); st.rerun()
                except Exception as e: st.error(f"Gagal: {e}")

# --- 7. SIDEBAR & DOWNLOAD EXCEL (DENGAN FIX LEBAR KOLOM) ---
st.sidebar.markdown("Simpan Rekap Kas Kecil")
if not df_raw.empty:
    all_kelompok = sorted([k for k in df_raw['Kelompok_Sheet'].unique() if k != ""], 
                          key=lambda x: (int(x.split(' ')[-1]), list(nama_bulan_id.values()).index(x.split(' ')[0])))
    
    def buat_excel(df_sel, grp_sel):
        wb = Workbook(); del wb['Sheet']
        thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        h_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
        summary_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
        title_fill = PatternFill(start_color="CC9900", end_color="CC9900", fill_type="solid")
        
        for p in grp_sel:
            ws = wb.create_sheet(title=p[:31])
            ws.merge_cells('A2:L3')
            ws['A2'] = "APLIKASI BIOS (BIAYA OPERASIONAL)"; ws['A2'].font = Font(bold=True, color="FFFFFF", size=14); ws['A2'].alignment = Alignment(horizontal="center", vertical="center"); ws['A2'].fill = title_fill
            headers = ["No", "URAIAN", "NAMA VENDOR", "POS MATA ANGGARAN", "GL ACCOUNT", "TANGGAL TRANSAKSI (SESUAI NOTA)", "JUMLAH PENGGUNAAN", "SETELAH PPN", "PPN", "PPH 23", "PPH 4(2)", "TOTAL"]
            ws.append([]); ws.append([]); ws.append(headers) 
            for col_num, val in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num); cell.fill = h_fill; cell.font = Font(bold=True); cell.border = thin; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            df_b = df_sel[df_sel['Kelompok_Sheet'] == p]
            curr_r = 6
            for i, r in enumerate(df_b.itertuples(), 1):
                t = "" if r.uraian == "Karcis Parkir Kendaraan Operasional" else pd.to_datetime(r.tanggal).strftime('%d/%m/%Y')
                ws.append([i, f"{i} {r.uraian}", r.vendor, "", "", t, r.jumlah, r.jumlah, "", "", "", r.jumlah])
                for col_idx in range(1, 13):
                    cell = ws.cell(row=curr_r, column=col_idx); cell.border = thin
                    if "Isi BBM" in r.uraian: cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    if "Karcis Parkir" in r.uraian: cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    if col_idx in [7, 8, 12]: cell.number_format = '#,##0'
                curr_r += 1
            
            # TOTAL BIRU
            ws.cell(row=curr_r, column=1).border = thin; ws.cell(row=curr_r, column=2).value = "TOTAL"; ws.cell(row=curr_r, column=2).font = Font(bold=True)
            ws.cell(row=curr_r, column=7).value = f"=SUM(G6:G{curr_r-1})"; ws.cell(row=curr_r, column=8).value = f"=SUM(H6:H{curr_r-1})"; ws.cell(row=curr_r, column=12).value = f"=SUM(L6:L{curr_r-1})"
            for col_idx in range(1, 13):
                cell = ws.cell(row=curr_r, column=col_idx); cell.fill = summary_fill; cell.font = Font(bold=True); cell.border = thin
                if col_idx in [7, 8, 12]: cell.number_format = '#,##0'
            curr_r += 1
            
            # RINGKASAN SISA
            labels = [("PENGAJUAN UANG MUKA", LIMIT_KAS), ("PENGGUNAAN UANG MUKA", f"=L{curr_r-1}"), ("SISA UANG MUKA", f"=L{curr_r}-L{curr_r+1}")]
            for offset, (lab, val) in enumerate(labels):
                row_idx = curr_r + offset
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=11)
                ws.cell(row=row_idx, column=1).value = lab; ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="right")
                target_cell = ws.cell(row=row_idx, column=12); target_cell.value = val
                for col_idx in range(1, 13):
                    c = ws.cell(row=row_idx, column=col_idx); c.fill = summary_fill; c.font = Font(bold=True); c.border = thin
                    if col_idx == 12: c.number_format = '#,##0'

            # FIX: LEBAR KOLOM OTOMATIS BIAR GAK MUNCUL #######
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 45 # Uraian
            ws.column_dimensions['C'].width = 20 # Vendor
            ws.column_dimensions['D'].width = 15 # Pos
            ws.column_dimensions['E'].width = 15 # GL
            ws.column_dimensions['F'].width = 25 # Tanggal
            ws.column_dimensions['G'].width = 18 # Jumlah Penggunaan
            ws.column_dimensions['H'].width = 18 # Setelah PPN
            ws.column_dimensions['I'].width = 12 # PPN
            ws.column_dimensions['J'].width = 12 # PPH 23
            ws.column_dimensions['K'].width = 12 # PPH 4(2)
            ws.column_dimensions['L'].width = 18 # TOTAL
            
        buf = BytesIO(); wb.save(buf); return buf.getvalue()

    if st.sidebar.button("💾 Siapkan Excel Keseluruhan"):
        nama_file_all = f"Rekap Kas Kecil 2026 {all_kelompok[-1]}.xlsx"
        st.sidebar.download_button("Download", buat_excel(df_raw, all_kelompok), nama_file_all)
    
    st.sidebar.divider()
    tabel_pil = st.sidebar.selectbox("Pilih Tabel:", all_kelompok)
    if st.sidebar.button(f"💾 Siapkan Excel {tabel_pil}"):
        nama_file_single = f"Rekap Kas Kecil 2026 {tabel_pil}.xlsx"
        st.sidebar.download_button(f"Download {tabel_pil}", buat_excel(df_raw, [tabel_pil]), nama_file_single)

st.sidebar.divider()
konf = st.sidebar.checkbox("Hapus Seluruh Data?")
if st.sidebar.button("Kosongkan Data", type="primary", disabled=not konf):
    conn.table("kas_kecil").delete().neq("id", 0).execute()
    st.rerun()
